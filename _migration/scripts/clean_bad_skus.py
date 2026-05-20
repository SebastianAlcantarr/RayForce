"""
clean_bad_skus.py
Elimina del CSV final:
  - Filas con SKU corrupto (contiene caracteres raros, como ABUã000X)
  - La fila de encabezado falsa (SKU = 'Codigo')
Regenera woocommerce_FINAL.csv y woocommerce_FINAL.xlsx
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

INPUT  = "woocommerce_FINAL.csv"
OUTPUT = "woocommerce_FINAL.csv"
EXCEL  = "woocommerce_FINAL.xlsx"

# SKUs a eliminar (corruptos + fila falsa de encabezado)
SKUS_ELIMINAR = {
    # ABUa = version corrupta de ABUn (ya existen los correctos ABUn000X)
    "ABU\u00e30001",        "ABU\u00e30001-BASE",
    "ABU\u00e30002",        "ABU\u00e30003",
    "ABU\u00e30004",        "ABU\u00e30005",
    "ABU\u00e30006",        "ABU\u00e30008",
    "ABU\u00e30009",
    # Fila de encabezado colada como dato
    "C\u00f3digo",
}

print(f"Leyendo {INPUT}...")
with open(INPUT, "r", encoding="utf-8-sig", errors="ignore") as f:
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames
    all_rows = list(reader)

print(f"  Filas antes: {len(all_rows)}")

# Filtrar filas malas
clean_rows = []
removed = []
for row in all_rows:
    sku = row.get("SKU", "").strip()
    if sku in SKUS_ELIMINAR:
        removed.append(sku)
    else:
        clean_rows.append(row)

print(f"  Filas eliminadas: {len(removed)}")
for s in removed:
    print(f"    - {s}")
print(f"  Filas despues: {len(clean_rows)}")

# Guardar CSV limpio
with open(OUTPUT, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(clean_rows)

print(f"\nCSV guardado: {OUTPUT}")

# Regenerar Excel
print(f"Regenerando {EXCEL}...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos"

hfill = PatternFill("solid", fgColor="1F4E79")
hfont = Font(color="FFFFFF", bold=True, size=10)

ws.append(fieldnames)
for ci in range(1, len(fieldnames) + 1):
    cell = ws.cell(row=1, column=ci)
    cell.fill = hfill
    cell.font = hfont

for row in clean_rows:
    ws.append([row.get(f, "") for f in fieldnames])

col_w = {"SKU": 16, "Type": 10, "Name": 50, "Description": 30,
         "Categories": 25, "Images": 60, "Regular price": 14, "Parent": 20}
for ci, cn in enumerate(fieldnames, 1):
    ws.column_dimensions[get_column_letter(ci)].width = col_w.get(cn, 12)

ws.freeze_panes = "A2"
wb.save(EXCEL)

print(f"Excel guardado: {EXCEL}")
print()
print("=" * 50)
print(f"LISTO. Filas finales: {len(clean_rows)}")
print("=" * 50)
