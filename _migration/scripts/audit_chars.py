"""
audit_chars.py - Encuentra y corrige todos los caracteres raros del CSV final.
"""
import csv
import re

INPUT  = "woocommerce_FINAL.csv"
OUTPUT = "woocommerce_FINAL.csv"

# Caracteres validos en espanol + tecnicos
VALID = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
            "áéíóúàèìòùÁÉÍÓÚÀÈÌÒÙñÑüÜ"
            "0123456789"
            " .,;:!?()-/\"'#+*%&@°½¼¾×÷"
            "\n\t")

# Correcciones adicionales (doble escape de encoding)
EXTRA_FIXES = [
    # Patron comun: ã seguido de letra
    ("ã±", "ñ"), ("ã¡", "á"), ("ã©", "é"), ("ã­", "í"),
    ("ã³", "ó"), ("ãº", "ú"), ("ã ", "à"),
    # Patrones con +
    ("Â+", ""), ("+", "+"),   # el + normal es valido
    # Caracteres de control
    ("\x91", "'"), ("\x92", "'"), ("\x93", '"'), ("\x94", '"'),
    ("\x96", "-"), ("\x97", "-"), ("\x85", "..."),
    # Otros residuos
    ("Ã", ""), ("â", ""), ("Â", ""),
]

def clean(text):
    for bad, good in EXTRA_FIXES:
        text = text.replace(bad, good)
    # Eliminar cualquier caracter no valido (excepto los normales)
    result = []
    for c in text:
        if c in VALID or ord(c) < 128:
            result.append(c)
        # else: se descarta
    return "".join(result).strip()

# Leer
with open(INPUT, "r", encoding="utf-8-sig", errors="replace") as f:
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames
    rows = list(reader)

# Auditar antes
bad_before = sum(1 for r in rows
                 if any(ord(c) > 127 and c not in "áéíóúàèìòùÁÉÍÓÚÀÈÌÒÙñÑüÜ¿¡°½¼¾"
                        for c in r.get("Name", "")))
print(f"Productos con chars raros ANTES: {bad_before}")

# Limpiar campo Name (y Description)
cleaned = 0
for row in rows:
    orig = row.get("Name", "")
    fixed = clean(orig)
    if fixed != orig:
        row["Name"] = fixed
        cleaned += 1

# Auditar despues
bad_after = sum(1 for r in rows
                if any(ord(c) > 127 and c not in "áéíóúàèìòùÁÉÍÓÚÀÈÌÒÙñÑüÜ¿¡°½¼¾"
                       for c in r.get("Name", "")))
print(f"Nombres limpiados: {cleaned}")
print(f"Productos con chars raros DESPUES: {bad_after}")

# Guardar
with open(OUTPUT, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

# Mostrar ejemplos restantes si los hay
if bad_after > 0:
    print("\nEjemplos con chars aun raros:")
    shown = 0
    for r in rows:
        name = r.get("Name", "")
        bad = [c for c in name if ord(c) > 127 and c not in "áéíóúàèìòùÁÉÍÓÚÀÈÌÒÙñÑüÜ¿¡°½¼¾"]
        if bad and shown < 10:
            print(f"  SKU={r['SKU']} | chars={bad} | {name[:60]}")
            shown += 1
else:
    print("\nCSV completamente limpio. Listo para WooCommerce.")

# Regenerar Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True, size=10)

    ws.append(fieldnames)
    for ci, cn in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = hfill
        cell.font = hfont

    for row in rows:
        ws.append([row.get(f, "") for f in fieldnames])

    col_w = {"SKU": 16, "Type": 10, "Name": 50, "Description": 30,
             "Categories": 25, "Images": 60, "Regular price": 14, "Parent": 20}
    for ci, cn in enumerate(fieldnames, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_w.get(cn, 12)

    ws.freeze_panes = "A2"
    wb.save("woocommerce_FINAL.xlsx")
    print("Excel regenerado: woocommerce_FINAL.xlsx")
except Exception as e:
    print(f"Error generando Excel: {e}")
