"""
merge_images.py
Toma las imagenes del CSV DEFINITIVO y las agrega al CSV con nombres mejorados.
Genera el Excel final listo para WooCommerce.
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFINITIVO_CSV = "woocommerce_DEFINITIVO (1).csv"
MEJORADO_CSV   = "woocommerce_nombres_mejorados.csv"
OUTPUT_EXCEL   = "woocommerce_FINAL.xlsx"
OUTPUT_CSV     = "woocommerce_FINAL.csv"

def main():
    # 1. Cargar imagenes del DEFINITIVO indexadas por SKU
    print(f"Leyendo imagenes de {DEFINITIVO_CSV}...")
    img_map = {}
    with open(DEFINITIVO_CSV, "r", encoding="utf-8-sig", errors="ignore") as f:
        for row in csv.DictReader(f):
            sku = row.get("SKU", "").strip()
            img = row.get("Images", "").strip()
            if sku and img:
                img_map[sku] = img

    real_imgs = sum(1 for v in img_map.values()
                    if v and "dummyimage" not in v and v != "")
    print(f"  -> {len(img_map)} SKUs con imagen | {real_imgs} imagenes reales (no placeholder)")

    # 2. Leer CSV mejorado y aplicar imagenes
    print(f"Leyendo {MEJORADO_CSV}...")
    with open(MEJORADO_CSV, "r", encoding="utf-8-sig", errors="ignore") as f:
        reader    = csv.DictReader(f)
        fieldnames = reader.fieldnames
        all_rows  = list(reader)

    print(f"  -> {len(all_rows)} filas")

    updated   = 0
    kept_dummy = 0

    for row in all_rows:
        sku = row.get("SKU", "").strip()
        if sku in img_map:
            new_img = img_map[sku]
            if new_img and "dummyimage" not in new_img:
                row["Images"] = new_img
                updated += 1
            else:
                kept_dummy += 1

    print(f"  -> Imagenes reales aplicadas  : {updated}")
    print(f"  -> Siguen con placeholder     : {kept_dummy}")

    # 3. Guardar CSV final (utf-8-sig para Excel)
    print(f"Guardando {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    # 4. Generar Excel final con formato
    print(f"Generando {OUTPUT_EXCEL}...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Estilo encabezado
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Escribir encabezados
    ws.append(fieldnames)
    for col_idx, col_name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align

    # Escribir datos
    for row in all_rows:
        ws.append([row.get(f, "") for f in fieldnames])

    # Ajustar anchos de columna
    col_widths = {
        "SKU": 16, "Type": 10, "Name": 50, "Description": 30,
        "Categories": 25, "Images": 60, "Regular price": 14,
        "Parent": 20,
    }
    for col_idx, col_name in enumerate(fieldnames, 1):
        width = col_widths.get(col_name, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Congelar primera fila
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_EXCEL)

    # Resumen
    total     = len(all_rows)
    con_img   = sum(1 for r in all_rows
                    if r.get("Images","") and "dummyimage" not in r.get("Images",""))
    sin_img   = total - con_img
    pct       = round(100 * con_img / total) if total else 0

    print()
    print("=" * 55)
    print("COMPLETADO")
    print(f"  Total productos           : {total}")
    print(f"  Con imagen real           : {con_img}  ({pct}%)")
    print(f"  Sin imagen (placeholder)  : {sin_img}")
    print(f"  CSV final                 : {OUTPUT_CSV}")
    print(f"  Excel final               : {OUTPUT_EXCEL}")
    print("=" * 55)


if __name__ == "__main__":
    main()
