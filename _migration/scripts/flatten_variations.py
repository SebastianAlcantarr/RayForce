"""
flatten_variations.py
Convierte el catalogo de WooCommerce:
  - Elimina todos los productos BASE (Type: variable)
  - Convierte variation -> simple
  - Limpia el campo Parent
  - Corrige el formato de Parent de id: a sku: (por si acaso)
  - Regenera CSV y Excel finales
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

INPUT  = "woocommerce_FINAL.csv"
OUTPUT = "woocommerce_FINAL.csv"
EXCEL  = "woocommerce_FINAL.xlsx"

print(f"Leyendo {INPUT}...")
with open(INPUT, "r", encoding="utf-8-sig", errors="ignore") as f:
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames
    all_rows = list(reader)

print(f"  Filas originales: {len(all_rows)}")

# Contar tipos antes
types_before = {}
for r in all_rows:
    t = r.get("Type", "").strip().lower()
    types_before[t] = types_before.get(t, 0) + 1
print(f"  Tipos antes: {types_before}")

# ── Transformacion ────────────────────────────────────────────────────────────
clean_rows = []
removed_variable = 0
converted_variation = 0

for row in all_rows:
    tipo = row.get("Type", "").strip().lower()

    # 1. Eliminar productos variable (los BASE)
    if tipo == "variable":
        removed_variable += 1
        continue

    # 2. Convertir variation -> simple
    if tipo == "variation":
        row["Type"] = "simple"
        # Limpiar campo Parent (ya no tiene padre)
        row["Parent"] = ""
        converted_variation += 1

    # 3. Si por alguna razon quedó un Parent con id:, convertirlo a sku:
    parent = row.get("Parent", "")
    if parent.startswith("id:"):
        row["Parent"] = "sku:" + parent[3:]

    clean_rows.append(row)

print()
print(f"  Eliminados (variable/BASE): {removed_variable}")
print(f"  Convertidos (variation->simple): {converted_variation}")
print(f"  Filas finales: {len(clean_rows)}")

# Contar tipos despues
types_after = {}
for r in clean_rows:
    t = r.get("Type", "").strip().lower()
    types_after[t] = types_after.get(t, 0) + 1
print(f"  Tipos despues: {types_after}")

# ── Guardar CSV ───────────────────────────────────────────────────────────────
print(f"\nGuardando {OUTPUT}...")
with open(OUTPUT, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(clean_rows)

# ── Regenerar Excel ───────────────────────────────────────────────────────────
print(f"Generando {EXCEL}...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos"

hfill = PatternFill("solid", fgColor="1F4E79")
hfont = Font(color="FFFFFF", bold=True, size=10)

ws.append(fieldnames)
for ci in range(1, len(fieldnames) + 1):
    ws.cell(row=1, column=ci).fill = hfill
    ws.cell(row=1, column=ci).font = hfont

for row in clean_rows:
    ws.append([row.get(f, "") for f in fieldnames])

col_w = {
    "SKU": 16, "Type": 10, "Name": 50, "Description": 30,
    "Categories": 25, "Images": 60, "Regular price": 14, "Parent": 20,
}
for ci, cn in enumerate(fieldnames, 1):
    ws.column_dimensions[get_column_letter(ci)].width = col_w.get(cn, 12)

ws.freeze_panes = "A2"
wb.save(EXCEL)

print()
print("=" * 55)
print("COMPLETADO")
print(f"  Productos simple  : {types_after.get('simple', 0)}")
print(f"  Productos variable: {types_after.get('variable', 0)}  (deberia ser 0)")
print(f"  CSV final         : {OUTPUT}")
print(f"  Excel final       : {EXCEL}")
print("=" * 55)
print()
print("NOTA: Para que los productos relacionados muestren chips")
print("entre si en la pagina, necesitaras un plugin de WooCommerce")
print("como 'WPC Linked Variation' o implementacion personalizada.")
